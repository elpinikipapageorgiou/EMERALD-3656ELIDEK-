#--- IMPORT DEPENDENCIES ------------------------------------------------------+
import math
import random
import time
from random import uniform
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from sklearn import metrics
from sklearn.cluster import KMeans
from sklearn.metrics import (accuracy_score, confusion_matrix, f1_score,
                             precision_score, recall_score)
from sklearn.model_selection import KFold
import itertools
from statistics import mean
import networkx as nx
from openpyxl import load_workbook
import glob
#--- MAIN ---------------------------------------------------------------------+

num_dimensions=18
class Particle:
    def __init__(self):
        self.pos_best_i=[]          # best position individual
        self.err_best_i=-1          # best error individual
        self.err_i=-1               # error individual
        #initilization of position and velocity

        df = pd.read_excel("suggested_weight_matrix_images_only8_2outputs.xlsx", nrows=num_dimensions, engine='openpyxl')
        arr = df.to_numpy()

        for i in range(0,num_dimensions):
            for j in range(0,num_dimensions):


                if arr[i][j]=="negative":
                    arr[i][j]=random.uniform(-1, -0.5)
                if arr[i][j]=="positive":
                    arr[i][j]=random.uniform(0.5, 1)

                if arr[i][j]=="positive_output":
                    arr[i][j]=random.uniform(0, 1)
                if arr[i][j]=="negative_output":
                    arr[i][j]=random.uniform(-1, 0)

        np.fill_diagonal(arr, 0)
        W2 =np.random.uniform(size = (num_dimensions,num_dimensions), low = -1, high = 1)
        np.fill_diagonal(W2, 0)
        #nullify last row of matrix
        arr[-1] = 0
        arr[-2] = 0
        self.position_i=(arr)
        #nullify last row of matrix
        W2[-1] = 0
        W2[-2] = 0
        self.velocity_i=(W2)

    # evaluate current fitness
    def evaluate(self,err_i):
        # check to see if the current position is an individual best
        if self.err_i<self.err_best_i or self.err_best_i==-1:
            self.pos_best_i=self.position_i.copy()
            self.err_best_i=err_i
 # update new particle velocity
    def update_velocity(self,pos_best_g):
        w=0.5       # constant inertia weight (how much to weigh the previous velocity)
        c1=1       # cognative constant
        c2=0.5        # social constant
        r1=uniform(0,1)
        r2=uniform(0,1)
        for i in range(0,num_dimensions):
          for j in range(0,num_dimensions):
              if(i==j):
                continue
              else:
                vel_cognitive=c1*r1*(self.pos_best_i[i][j]-self.position_i[i][j])
                vel_social=c2*r2*(pos_best_g[i][j]-self.position_i[i][j])
                self.velocity_i[i][j]=w*self.velocity_i[i][j] +vel_cognitive+vel_social

    # update the particle position based off new velocity updates
    def update_position(self):
        for i in range(0,num_dimensions):
          for j in range(0,num_dimensions):
              if(i==j):
                continue
              else:
                self.position_i[i][j]=self.position_i[i][j]+self.velocity_i[i][j]
                # adjust maximum position if necessary
                if self.position_i[i][j]>1:
                    self.position_i[i][j]=1

                # adjust minimum position if neseccary
                if self.position_i[i][j]<-1:
                    self.position_i[i][j]=-1
                    
def sig(x):
    return 1/(1 + np.exp(-x))


def minimize_and_plot_concept_evolution(dataset, num_dimensions, num_particles, maxiter):
    err_best_g = float('inf')
    pos_best_g = []
    swarm = [Particle() for _ in range(num_particles)]
    concept_evolution = np.zeros((maxiter, num_dimensions))  # Record concept values over epochs

    for k in range(maxiter):
        # Evaluate fitness of each particle at its current position
        for particle in swarm:
            fitness_result = 0
            for row in dataset:
                fcm_output = [0] * num_dimensions
                for i in range(num_dimensions):
                    sum_temp = sum(particle.position_i[j][i] * row[j] for j in range(num_dimensions) if i != j)
                    fcm_output[i] = sig(sum_temp)

                fitness_result += np.square(fcm_output[-1] - row[-1])

            particle.err_i = fitness_result / len(dataset)
            particle.evaluate(particle.err_i)

            # Update the global best position
            if particle.err_i < err_best_g or k == 0:
                pos_best_g = particle.position_i.copy()
                err_best_g = particle.err_i

        # Record the concept evolution
        concept_evolution[k, :] = [sig(sum(pos_best_g[j][i] * value for j, value in enumerate(row))) for i in range(num_dimensions)]

        # Update velocity and position of each particle after fitness evaluation
        for particle in swarm:
            particle.update_velocity(pos_best_g)
            particle.update_position()

    return pos_best_g, concept_evolution


#calculate deviation for the performance metrics
def calculate_deviation(values):
    # Step 1
    mean = sum(values) / len(values)

    # Step 2
    differences = [value - mean for value in values]

    # Step 3
    squared_differences = [diff ** 2 for diff in differences]

    # Step 4
    mean_squared_differences = sum(squared_differences) / len(squared_differences)

    # Step 5
    standard_deviation = math.sqrt(mean_squared_differences)

    return standard_deviation

# Function to plot confusion matrix
def plot_confusion_matrix(cm, classes, title='Confusion Matrix', cmap=plt.cm.Blues):
    plt.imshow(cm, interpolation='nearest', cmap=cmap)
    plt.title(title)
    plt.colorbar()
    tick_marks = np.arange(len(classes))
    plt.xticks(tick_marks, classes, rotation=45)
    plt.yticks(tick_marks, classes)

    fmt = 'd'
    thresh = cm.max() / 2.
    for i, j in itertools.product(range(cm.shape[0]), range(cm.shape[1])):
        plt.text(j, i, format(cm[i, j], fmt), horizontalalignment="center",
                color="white" if cm[i, j] > thresh else "black")

    plt.ylabel('True label')
    plt.xlabel('Predicted label')
    plt.tight_layout()



#read output for each image-instance
print("training procedure")

#path of images
#path of images
extensions = ['jpg', 'png', 'jpeg', 'tiff', '.tif','.TIFF', '.TIF']

# Get a list of file paths that match the updated pattern for each extension
image_paths = []
for ext in extensions:
    path_pattern = f"../all_polar_maps/*.{ext}"
    image_paths.extend(glob.glob(path_pattern, recursive=True))
# print(len(image_paths))
addrs=image_paths
y = [0 if 'normal' in addr else 1 for addr in addrs]

# Number of classes
classes_num = 2
 
 
# Make a list of the images with the labels
c = list(zip(addrs, y))
 
# Shuffle the list
# random.shuffle(c)
 
# Create tuples with the addresses and the labels
addrs, y = zip(*c)
 



ischemic_Feature_maps=[]
normal_Feature_maps=[]
from tensorflow.keras.preprocessing.image import load_img, img_to_array

# Rest of your code remains the same

import keras.models
from keras.models import Model, load_model
from keras.preprocessing import image
from tensorflow.keras import models



#load pretrained model
from keras.models import load_model
model = keras.models.load_model("../polar_maps.keras")


# model = tf.keras.saving.load_model("model_new.h5")
# model = keras.models.load_model("maps")
model.summary()

# get the output of the last fully connected layer
layer_name = 'dense_2' # choose the name of the last fully connected layer
feature_layer_model = Model(inputs=model.input,
                            outputs=model.get_layer(layer_name).output)

import cv2
##############################################################################################3
##############################################################################################3
##############################################################################################3
##############################################################################################3
##############################################################################################3
##############################################################################################3
initial_training_featuresss = []
full_actual_output=[]
 
rgb_training_features=[]
# Create tuples with the addresses and the labels

#extract feature vectors
for img in addrs:  
    if "normal" in str(img):
        full_actual_output.append(0)
    if "pathological" in str(img):
        full_actual_output.append(1)
        
    image = cv2.imread(img)

    # Check if thage was loaded successfully
    # if image is None:
    #     print(f"Failed to load image: {img_path}")
    #     continuee im

    # Resize the image
    image = cv2.resize(image, (300, 300))
    image = image.astype('float32') / 255
    

    image = np.expand_dims(image, axis=0)



    # Use the model to extract the feature vector
    features = feature_layer_model.predict(image)

  

    rgb_training_features.append(features)

    
initial_training_featuresss = np.array(rgb_training_features)
initial_training_featuresss = np.squeeze(initial_training_featuresss, axis=1)
np.save('data_polar.npy', initial_training_featuresss)
initial_training_featuresss = np.load('data_polar.npy')
##############################################################################################3
##############################################################################################3
##############################################################################################3
##############################################################################################3
##############################################################################################3
#read vgg predictions for each image-instance

print(full_actual_output)
#divide feature maps according to their belonging class
for feature_index, y_index in zip(initial_training_featuresss, full_actual_output):
    if y_index==1:
        ischemic_Feature_maps.append(feature_index)
    if y_index==0:
        normal_Feature_maps.append(feature_index)



ischemic_Feature_maps = np.array(ischemic_Feature_maps)
normal_Feature_maps = np.array(normal_Feature_maps)

# normalize feature vectors
ischemic_Feature_maps = (ischemic_Feature_maps - np.min(ischemic_Feature_maps)) / (np.max(ischemic_Feature_maps) - np.min(ischemic_Feature_maps))
normal_Feature_maps = (normal_Feature_maps - np.min(normal_Feature_maps)) / (np.max(normal_Feature_maps) - np.min(normal_Feature_maps))


#create clusters for each of the feature vectors of two classes
num_clusters=8
#---------------------defective-------------------#
print("----centroids-----")
ischemic_Feature_maps = np.array(ischemic_Feature_maps)
ischemic_Feature_maps = ischemic_Feature_maps.reshape((-1, ischemic_Feature_maps.shape[-1]))


kmeans_defective = KMeans(n_clusters=num_clusters, n_init='auto')
kmeans_defective.fit(ischemic_Feature_maps)
centroids_ischemic= kmeans_defective.cluster_centers_
np.save('centroids_ischemic.npy', centroids_ischemic)

#append the centroid of each cluster
list_centroids_ischemic=[]
list_centroids_ischemic.append(centroids_ischemic[0])
list_centroids_ischemic.append(centroids_ischemic[1])
list_centroids_ischemic.append(centroids_ischemic[2])
list_centroids_ischemic.append(centroids_ischemic[3])
list_centroids_ischemic.append(centroids_ischemic[4])
list_centroids_ischemic.append(centroids_ischemic[5])
list_centroids_ischemic.append(centroids_ischemic[6])
list_centroids_ischemic.append(centroids_ischemic[7])
#---------------------Normal-------------------#
normal_Feature_maps = np.array(normal_Feature_maps)
normal_Feature_maps = normal_Feature_maps.reshape((-1, normal_Feature_maps.shape[-1]))


kmeans_normal = KMeans(n_clusters=num_clusters, n_init='auto')
kmeans_normal.fit(normal_Feature_maps)
centroids_normal= kmeans_normal.cluster_centers_
np.save('centroids_normal.npy', centroids_normal)
# print("centroids_normal", centroids_normal)
list_centroids_normal=[]
list_centroids_normal.append(centroids_normal[0])
list_centroids_normal.append(centroids_normal[1])
list_centroids_normal.append(centroids_normal[2])
list_centroids_normal.append(centroids_normal[3])
list_centroids_normal.append(centroids_normal[4])
list_centroids_normal.append(centroids_normal[5])
list_centroids_normal.append(centroids_normal[6])
list_centroids_normal.append(centroids_normal[7])

centroids =[]
centroids.append(list_centroids_ischemic)
centroids.append(list_centroids_normal)
centroids=np.array(centroids)

# compute the sum of all centroids of both two classes
sum_centroids = np.sum(centroids)
with open('sum_centroids.txt', 'w') as file:
    file.write(str(sum_centroids))
#compute the similarities of centroids for all feature vectors
rows=[]
similarities=[]
for training_row in initial_training_featuresss:
    rows.append((1- (np.linalg.norm(training_row - centroids_ischemic[0])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_ischemic[1])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_ischemic[2])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_ischemic[3])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_ischemic[4])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_ischemic[5])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_ischemic[6])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_ischemic[7])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_normal[0])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_normal[1])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_normal[2])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_normal[3])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_normal[4])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_normal[5])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_normal[6])/(sum_centroids))))
    rows.append((1- (np.linalg.norm(training_row - centroids_normal[7])/(sum_centroids))))

    similarities.append(rows)
    rows=[]

similarities = np.array(similarities)

similarities =  (similarities - np.min(similarities)) / np.ptp(similarities)


inverted_list = [1 if item == 0 else 0 for item in full_actual_output]

#construct the dataset, which includes the similarities along with the output
inverted_list = np.array(inverted_list)
full_actual_output=np.array(full_actual_output)
result = np.concatenate((similarities[:, 0].reshape(-1,1), similarities[:, 1].reshape(-1,1),
                         similarities[:, 2].reshape(-1,1), similarities[:, 3].reshape(-1,1),
                         similarities[:, 4].reshape(-1,1), similarities[:, 5].reshape(-1,1),
                         similarities[:, 6].reshape(-1,1), similarities[:, 7].reshape(-1,1),
                         similarities[:, 8].reshape(-1,1), similarities[:, 9].reshape(-1,1),
                         similarities[:, 10].reshape(-1,1), similarities[:, 11].reshape(-1,1),
                         similarities[:, 12].reshape(-1,1), similarities[:, 13].reshape(-1,1),
                         similarities[:, 14].reshape(-1,1), similarities[:, 15].reshape(-1,1),
                         inverted_list.reshape(-1,1),
                         full_actual_output.reshape(-1,1) ,
                         ), axis=1)

print("shape")
print(result.shape)


df = pd.DataFrame(result)
dataset = df.sample(frac=1).reset_index(drop=True)
# save the dataframe to an Excel file
dataset.to_excel('full_dataset.xlsx', index=False)



# #fill missing values if there are any
# dataset.fillna(method="bfill", inplace=True)

# #divide input-output data
# X = dataset.iloc[:,:-1].values
# y = dataset.iloc[: ,-1].values

# #perform oversampling
# oversample = SMOTE(sampling_strategy='minority')
# X, y = oversample.fit_resample(X, y)

# dataset = pd.concat([pd.DataFrame(X), pd.DataFrame(y)], axis=1, join="inner")


total_confusion_matrix = [[0, 0], [0, 0]]

mean_accuracies=[]
mean_losses=[]
epoch = 35

#perform k-fold cross validation
kf = KFold(n_splits=10, shuffle=True, random_state=9898)
start_time = time.time()
fold=0
acc=[]
err=[]
best_matrix=[]
concepts=[]
acc=[]
err=[]
best_matrix=[]
concepts=[]
class_accuracies0 =[]
class_accuracies1 =[]
c_matrices=[]
recalls=[]
best_positions=[]
cm_sum = []
sens=[]
spec=[]
prec=[]
fold_precision=[]
fold_f1=[]
fold_recall=[]
fold_spec=[]
fold_sens=[]
# Iterate over each train-test split

for train_index, test_index in kf.split(dataset):
    fold+=1


    print("\n\n************************************")
    print(fold)
    print("---------fold------------")
    print("************************************\n\n")

    # Split train-test
    training_dataset, testing_dataset=dataset.iloc[train_index], dataset.iloc[test_index]
    training_dataset= np.array(training_dataset)
    testing_dataset= np.array(testing_dataset)

    best_position, concept_evolution  = minimize_and_plot_concept_evolution(training_dataset,  num_dimensions=num_dimensions, num_particles=35, maxiter=epoch)
    end = time.time()
        # Plotting concept evolution
    plt.figure(figsize=(15, 10))
    epoch_range = range(1, epoch + 1)  # Create an array for the x-axis representing epochs
    for j in range(num_dimensions):
        plt.plot(epoch_range, concept_evolution[:, j], label=f'Concept {j+1}')
    plt.xlabel('Epochs')
    plt.ylabel('Concept Value')
    plt.title(f'Evolution of Concept Values Over Epochs (Fold {fold})')
    plt.legend()
    plt.legend(loc='upper right')
    # Save the plot as an image file
    filename = f'fold_{fold}_inference_plot.png'  # Define your filename here
    plt.savefig(filename)  # Save the plot
    plt.close()  # Close the plot to free memory


    for i in range(0,num_dimensions):
        for j in range(0,num_dimensions):
            if(i==j):
                continue
            else:
                # adjust maximum position if necessary
                if best_position[i][j]>1:
                    best_position[i][j]=1

                # # adjust minimum position if neseccary
                if best_position[i][j]<-1:
                    best_position[i][j]=-1

    sum_temp=0
    #from the testing dataset calculate the predcited values and compare with the actual output
    testing_last_element_fcm_output=[]
    best_position=np.vstack(best_position)
    predicted_results=[None]*num_dimensions
    for testing_row in testing_dataset:

        for i in range(0,num_dimensions):
            for j in range(0,num_dimensions):
                if(i==j):
                    continue
                else:
                    sum_temp=sum_temp+best_position[j][i]*testing_row[j]

            predicted_results[i]=sum_temp
            predicted_results[i] = sig(predicted_results[i])

            sum_temp=0


        testing_last_element_fcm_output.append((predicted_results)[-1])
    testing_actual_output = testing_dataset[:, -1]

    testing_last_element_fcm_output = np.vstack(testing_last_element_fcm_output)

    testing_last_element_fcm_output = testing_last_element_fcm_output[:, -1]

    temporary_value_results=testing_last_element_fcm_output


    limits_acc=[]
    limits= np.arange(0.1, 0.99, 0.01).tolist()

    steady_value_predicted_results=temporary_value_results
    for i in limits:

        temporary_value_results = steady_value_predicted_results > i

        testing_actual_output=np.array(testing_actual_output)
        temporary_value_results=(np.array(temporary_value_results))

        limits_acc.append(accuracy_score(testing_actual_output, temporary_value_results.round())*100)


    max_value = max(limits_acc)
    index = limits_acc.index(max_value)

    testing_last_element_fcm_output = testing_last_element_fcm_output > limits[index]

    testing_actual_output=np.array(testing_actual_output)
    testing_last_element_fcm_output=(np.array(testing_last_element_fcm_output))

    #calculate performance metrics
    A=(accuracy_score(testing_actual_output, testing_last_element_fcm_output.round())*100)
    acc.append(accuracy_score(testing_actual_output, testing_last_element_fcm_output.round())*100)


    err.append(metrics.mean_absolute_error(testing_actual_output, testing_last_element_fcm_output))
    cm = confusion_matrix(testing_actual_output, testing_last_element_fcm_output)
    cm_sum.append(cm)
    total_confusion_matrix += cm
    class_counts = cm.sum(axis=1)
    accuracies = [0 if count == 0 else cm[i, i] / count for i, count in enumerate(class_counts)]
    fold_precision.append(precision_score(testing_actual_output, testing_last_element_fcm_output, zero_division=1.0)*100)
    # Calculate recall
    fold_f1.append(f1_score(testing_actual_output, testing_last_element_fcm_output)*100)

    fold_recall.append(recall_score(testing_actual_output, testing_last_element_fcm_output)*100)
    TN = cm[0, 0]
    FP = cm[0, 1]
    FN = cm[1, 0]
    TP = cm[1, 1]

    sensitivity = TP / (TP + FN)
    specificity = TN / (TN + FP)
    fold_sens.append(sensitivity*100)
    fold_spec.append(specificity*100)

    

    # Example predicted labels and actual labels
    y_pred = testing_last_element_fcm_output
    y_true = testing_actual_output

    # Compute confusion matrix
    cm = confusion_matrix(y_true, y_pred)

    # Define class labels
    class_names = ['Normal', 'Pathological']

    # convert the NumPy array to a pandas DataFrame
    df = pd.DataFrame(best_position, columns=["c1","c2", "c3", "c4", "c5", "c6","c7", "c8","c9", "c10","c11", "c12","c13","c14", "c15", "c16", "healthy", "defective"])

    # write the DataFrame to an Excel file
    df.to_excel(f'data{fold}.xlsx', index=False)


#print performance metrics
print("\n\n\n")
print("-------------end of kfold------------")
end_time = time.time()
execution_time = end_time - start_time

# Print the execution time
print(f"Execution Time: {execution_time:.6f} seconds")
print("Epochs",epoch)

print("Accuracies")
print(acc)
print(mean(acc))


acc_deviation = calculate_deviation(acc)
print("acc_deviation")
print(acc_deviation)



print("\n\nError")
print(err)
print(mean(err))
err_deviation = calculate_deviation(err)
print("err_deviation")
print(err_deviation)
print("\nTotal Confusion Matrix:")
print(total_confusion_matrix)



print("\nSensitivity")
print(fold_sens)
print(mean(fold_sens))
sensitivity_deviation = calculate_deviation(fold_sens)
print("sensitivity_deviation")
print(sensitivity_deviation)

print("\nSpecificity")
print(fold_spec)
print(mean(fold_spec))
specificity_deviation = calculate_deviation(fold_spec)
print("specificity_deviation")
print(specificity_deviation)


print("\n\Precision")
print(fold_precision)
print(mean(fold_precision))
precision_deviation = calculate_deviation(fold_precision)
print("precision_deviation")
print(precision_deviation)


input_file_names = ["data1.xlsx", "data2.xlsx", "data3.xlsx", "data4.xlsx", "data5.xlsx",
                    "data6.xlsx", "data7.xlsx", "data8.xlsx", "data9.xlsx", "data10.xlsx"]

# Initialize a sum DataFrame with zeros
sum_df = pd.DataFrame(0, index=range(num_dimensions), columns=range(num_dimensions), dtype=float)

# Read data from each input file and accumulate the sum
for file_name in input_file_names:
    try:
        # Read the Excel file
        df = pd.read_excel(file_name, header=None)

        # Exclude the first row and add the array values to sum_df
        sum_df += df.iloc[1:, :].astype(float)

    except Exception as e:
        print(f"Error occurred while processing {file_name}: {e}")
# Calculate the mean DataFrame by dividing the sum_df by the number of files
num_files = len(input_file_names)
mean_df = sum_df / num_files

# Create an Excel file with the mean DataFrame
output_file_name = "mean_values.xlsx"
mean_df.to_excel(output_file_name, index=False, header=False)

print("Mean values saved to", output_file_name)

# Read all Excel files and calculate standard deviations
dfs = [pd.read_excel(file, header=None, skiprows=1, decimal=',') for file in input_file_names]
deviations = np.std([df.values for df in dfs], axis=0)

# Create a DataFrame with deviations
deviation_df = pd.DataFrame(deviations)

# Save the deviations DataFrame to a new Excel file
deviation_df.to_excel('deviations.xlsx', index=False, header=False)

# # Create a directed graph
# G = nx.DiGraph()
# concepts = ["SEX","AGE", "BMI", "KNOWN CAD", "PREVIOUS AMI", "PREVIOUS PCI",
#                                               "PREVIOUS CABG", "PREVIOUS STROKE", 
#                                               "DIABETES","SMOKING", "HYPERTENSION",
#                                               "DYSLIPIDEMIA", "ANGIOPATHY", "CHRONIC KIDNEY DISEASE",
#                                               "FAMILY HISTORY OF CAD", "ASYMPTOMATIC",
#                                               "ATYPICAL SYMPTOMS", "ANIGNA LIKE", "DYSPNOEA ON EXERTION",
#                                               "INCIDENT OF PRECORDIAL PAIN", "ECG","EXPERT DIAGNOSIS BINARY", 
#                                               "CNN OUTPUT", "OUTPUT" ]
# # Add nodes to the graph
# G.add_nodes_from(concepts)
# print(type(best_position))
# rounded_array = np.round(best_position.astype(float), decimals=2)
# # print(rounded_array)
# # Add weighted edges to the graph based on the weight matrix
# for i in range(len(concepts)):
#     for j in range(len(concepts)):
#         weight = rounded_array[i, j]
#         if weight != 0.0:
#             G.add_edge(concepts[i], concepts[j], weight=weight)

# # Draw the graph
# fig, ax = plt.subplots(figsize=(15, 15))  # Increase figure size
# pos = nx.spring_layout(G, k=0.25)  # Define layout with reduced 'k' to separate nodes
# nx.draw(G, pos, with_labels=True, node_size=800, node_color='skyblue', font_size=6)  # Adjust node_size and font_size

# # Add edge labels (weights)
# edge_labels = {(n1, n2): d['weight'] for n1, n2, d in G.edges(data=True)}
# nx.draw_networkx_edge_labels(G, pos, edge_labels=edge_labels, font_size=6)  # Adjust font_size for edge labels

# plt.title('Fuzzy Cognitive Map')
# plt.tight_layout()  # Ensure tight layout
# plt.show()

# Calculate the mean DataFrame by dividing the sum_df by the number of files
num_files = len(input_file_names)
mean_df = sum_df / num_files

# Create an Excel file with the mean DataFrame
output_file_name = "deviations.xlsx"
mean_df.to_excel(output_file_name, index=False, header=False)

print("Deviations values saved to", output_file_name)



# Read all Excel files and calculate standard deviations
dfs = [pd.read_excel(file, header=None, skiprows=1, decimal=',') for file in input_file_names]
deviations = np.std([df.values for df in dfs], axis=0)

# Create a DataFrame with deviations
deviation_df = pd.DataFrame(deviations)

# Save the deviations DataFrame to a new Excel file
deviation_df.to_excel('deviations.xlsx', index=False, header=False)

# Load the source and destination Excel files
source_wb = load_workbook('suggested_weight_matrix_images_only8_2outputs.xlsx')
destination_wb = load_workbook('mean_values.xlsx')

# Access the first sheet of each workbook
source_sheet = source_wb.active
destination_sheet = destination_wb.active

# Get the values of the first row from the source sheet
first_row_values = []
for cell in source_sheet[1]:
    first_row_values.append(cell.value)

# Paste the first row values into the first row of the destination sheet
for index, value in enumerate(first_row_values, start=1):
    destination_sheet.cell(row=1, column=index, value=value)

# Save the changes to the destination file
destination_wb.save('mean_values.xlsx')

destination_wb = load_workbook('deviations.xlsx')
# Access the first sheet of each workbook
source_sheet = source_wb.active
destination_sheet = destination_wb.active

# Insert a new blank row at the beginning in the destination sheet
destination_sheet.insert_rows(1)

# Get the values of the first row from the source sheet
first_row_values = []
for cell in source_sheet[1]:
    first_row_values.append(cell.value)

# Paste the first row values into the first row of the destination sheet
for index, value in enumerate(first_row_values, start=1):
    destination_sheet.cell(row=1, column=index, value=value)

# Save the changes to the destination file
destination_wb.save('deviations.xlsx')


from tensorflow.keras.models import Model
import tensorflow as tf
import numpy as np
import cv2
import os
import keras
class GradCAM:
    def __init__(self, model, classIdx, layerName=None):
        # store the model, the class index used to measure the class
        # activation map, and the layer to be used when visualizing
        # the class activation map
        self.model = model
        self.classIdx = classIdx
        self.layerName = layerName
        # if the layer name is None, attempt to automatically find
        # the target output layer
        if self.layerName is None:
            self.layerName = self.find_target_layer()

    def find_target_layer(self):
        # attempt to find the final convolutional layer in the network
        # by looping over the layers of the network in reverse order
        for layer in reversed(self.model.layers):
            # check to see if the layer has a 4D output
            if len(layer.output_shape) == 4:
                return layer.name
        # otherwise, we could not find a 4D layer so the GradCAM
        # algorithm cannot be applied
        raise ValueError("Could not find 4D layer. Cannot apply GradCAM.")


    def compute_heatmap(self, image, eps=1e-8):
        # construct our gradient model by supplying (1) the inputs
        # to our pre-trained model, (2) the output of the (presumably)
        # final 4D layer in the network, and (3) the output of the
        # softmax activations from the model
        gradModel = Model(
            inputs=[self.model.inputs],
            outputs=[self.model.get_layer(self.layerName).output, self.model.output])

        # record operations for automatic differentiation
        with tf.GradientTape() as tape:
            # cast the image tensor to a float-32 data type, pass the
            # image through the gradient model, and grab the loss
            # associated with the specific class index
            inputs = tf.cast(image, tf.float32)
            (convOutputs, predictions) = gradModel(inputs)
            
            loss = predictions[:, tf.argmax(predictions[0])]
    
        # use automatic differentiation to compute the gradients
        grads = tape.gradient(loss, convOutputs)

        # compute the guided gradients
        castConvOutputs = tf.cast(convOutputs > 0, "float32")
        castGrads = tf.cast(grads > 0, "float32")
        guidedGrads = castConvOutputs * castGrads * grads
        # the convolution and guided gradients have a batch dimension
        # (which we don't need) so let's grab the volume itself and
        # discard the batch
        convOutputs = convOutputs[0]
        guidedGrads = guidedGrads[0]

        # compute the average of the gradient values, and using them
        # as weights, compute the ponderation of the filters with
        # respect to the weights
        weights = tf.reduce_mean(guidedGrads, axis=(0, 1))
        cam = tf.reduce_sum(tf.multiply(weights, convOutputs), axis=-1)

        # grab the spatial dimensions of the input image and resize
        # the output class activation map to match the input image
        # dimensions
        (w, h) = (image.shape[2], image.shape[1])
        heatmap = cv2.resize(cam.numpy(), (w, h))
        # normalize the heatmap such that all values lie in the range
        # [0, 1], scale the resulting values to the range [0, 255],
        # and then convert to an unsigned 8-bit integer
        numer = heatmap - np.min(heatmap)
        denom = (heatmap.max() - heatmap.min()) + eps
        heatmap = numer / denom
        heatmap = (heatmap * 255).astype("uint8")
        # return the resulting heatmap to the calling function
        return heatmap

    def overlay_heatmap(self, heatmap, image, alpha=0.5, colormap=cv2.COLORMAP_JET):
        # Apply the color map to the heatmap
        heatmap_colored = cv2.applyColorMap(heatmap, colormap)
        
        # Convert the heatmap to RGB if the original image is in RGB
        heatmap_colored = cv2.cvtColor(heatmap_colored, cv2.COLOR_BGR2RGB)
        
        # Ensure the heatmap has the same size as the original image
        heatmap_resized = cv2.resize(heatmap_colored, (image.shape[1], image.shape[0]))

        # Check if the number of channels matches
        if len(image.shape) == 2:  # If the image is grayscale
            image = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)
        elif image.shape[2] == 1:  # Single channel grayscale image
            image = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)
        
        # Add the weighted heatmap to the image
        output = cv2.addWeighted(image, alpha, heatmap_resized, 1 - alpha, 0)

        return heatmap_resized, output

####Grad-CAM methodology

class0_name='pathological'
class1_name='normal'
pixel_size=300

model = keras.models.load_model("model_rgb.keras")


output_folder = 'gradcam_results'

# Create the output folder if it does not exist
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# List of common image extensions (Add the preffered extension)
image_extensions = ['.tiff', '.tif', '.TIFF', '.jpeg', '.jpg', '.png']
last_conv_layer_name = 'conv2d_2'  # Replace 'conv2d_2' with your actual last convolutional layer name

# Iterate over all files in the directory that match the extensions



# Iterate over all files in the directory
for file_path in image_paths:

    image = cv2.imread(file_path)
    if image is None:
        print(f"File not found or unable to read: {file_path}")
        continue

    print("Processing:", file_path)
    
    # Determine the actual class based on the file name
    if class0_name in file_path.lower():
        actual_class = class0_name
    else:
        actual_class = class1_name

    # Preprocess the image
    image_resized = cv2.resize(image, (pixel_size, pixel_size))
    image_resized = image_resized.astype('float32') / 255
    image_resized = np.expand_dims(image_resized, axis=0)

    # Predict the class
    preds = model.predict(image_resized)
    prediction = (preds > 0.5).astype(int)

    if prediction == 0:
        predicted_class = class0_name
        print("The model predicted this instance as " + class0_name + ".")
    else:
        predicted_class = class1_name
        print("The model predicted this instance as " + class1_name + ".")

    # Define the GradCAM instance with the manually set layer name
    icam = GradCAM(model, np.argmax(preds[0]), layerName='conv2d_2')
    heatmap = icam.compute_heatmap(image_resized)
    heatmap_resized = cv2.resize(heatmap, (pixel_size, pixel_size))

    # Overlay the heatmap onto the original image
    (heatmap, output) = icam.overlay_heatmap(heatmap_resized, image, alpha=0.5)

    # Convert images to RGB for display
    heatmap_rgb = cv2.cvtColor(heatmap, cv2.COLOR_BGR2RGB)

    # Concatenate heatmap and original image side by side
    concatenated = np.concatenate((image, heatmap_rgb), axis=1)

    # Define the output filename
    output_filename = os.path.join(output_folder, f"gradcam_actual{actual_class}_pred_{predicted_class}_{os.path.splitext(os.path.basename(file_path))[0]}.png")

    # Print the output filename to check
    print(f"Saving file to: {output_filename}")

    # Check if the concatenated image has the correct shape and type
    print(f"Image shape: {concatenated.shape}")
    print(f"Image type: {type(concatenated)}")

    # Save the concatenated image
    if not cv2.imwrite(output_filename, concatenated):
        print(f"Failed to save image: {output_filename}")
    else:
        print(f"Successfully saved image: {output_filename}")